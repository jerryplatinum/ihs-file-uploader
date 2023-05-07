import csv
import os
import shutil
from datetime import datetime
import time
from openpyxl import load_workbook

def retry_decorator(retry_count, retry_wait):
    def decorator(func):
        def wrapper(*args, **kwargs):
            for i in range(retry_count):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    if i < retry_count - 1:
                        print(f"Error: {e}. Retrying in {retry_wait} minutes...")
                        time.sleep(retry_wait * 60)
                    else:
                        raise e
        return wrapper
    return decorator

@retry_decorator(2, 5)
def compare_files(src_path, template_path, log_path, quarantined_path, prod_path, archive_path):
    files_to_process = []
    
    # Step 1: Get all files in the source directory
    for root, _, files in os.walk(src_path):
        for file in files:
            if file.endswith(('.csv', '.xls', '.xlsx')):
                files_to_process.append(os.path.join(root, file))
                
    summary_report = []
    
    for file_path in files_to_process:
        file_status = 'Success'
        file_error = ''
        
        # Step 2: Find the corresponding template file
        containing_folder = os.path.dirname(file_path).split(os.sep)[-1]
        template_folder = os.path.join(template_path, containing_folder)
        
        if not os.path.exists(template_folder):
            print(f"Error: Template folder not found - {template_folder}")
            continue
        
        template_files = [f for f in os.listdir(template_folder) if f.endswith(('.csv', '.xls', '.xlsx'))]
        
        if not template_files:
            print(f"Error: No template files found in {template_folder}")
            continue
        
        template_file = None
        for temp_file in template_files:
            if temp_file.split('.')[0].lower() in file_path.lower():
                template_file = os.path.join(template_folder, temp_file)
                break
        
        if not template_file:
            print(f"Error: No matching template file found for {file_path}")
            continue
        
        # Step 3: Compare and validate the contents
        if file_path.endswith('.csv'):
            header_mismatch = compare_csv_headers(file_path, template_file)
            if header_mismatch:
                file_status = 'Error'
                file_error = f"Header Mismatch: {header_mismatch}"
                quarantine_file(file_path, quarantined_path, file_error)
        else:
            header_mismatch, sheet_mismatch = compare_excel_headers(file_path, template_file)
            if header_mismatch or sheet_mismatch:
                file_status = 'Error'
                file_error = f"Header Mismatch: {header_mismatch}; Sheet Mismatch: {sheet_mismatch}"
                quarantine_file(file_path, quarantined_path, file_error)
                
        # Step 4: Create the log entry
        log_entry = create_log_entry(file_path, file_status, file_error)
        summary_report.append(log_entry)
        
        # Step 5: Copy the successful files
        if file_status == 'Success':
            copy_successful_file(file_path, template_file, prod_path, archive_path)
            
    # Step 6: Generate the summary report
    generate_summary_report(summary_report, log_path)

def compare_csv_headers(file_path, template_file):
       with open(file_path, newline='') as f:
        file_headers = list(csv.reader(f))[0]
        with open(template_file, newline='') as f:
            template_headers = list(csv.reader(f))[0]
            
        return [h for h in file_headers if h not in template_headers]

def compare_excel_headers(file_path, template_file):
    wb = load_workbook(filename=file_path, read_only=True)
    template_wb = load_workbook(filename=template_file, read_only=True)
    
    header_mismatch = {}
    sheet_mismatch = [sheet for sheet in wb.sheetnames if sheet not in template_wb.sheetnames]
    
    for sheet in wb.sheetnames:
        if sheet not in template_wb.sheetnames:
            continue
            
        file_ws = wb[sheet]
        template_ws = template_wb[sheet]
        file_headers = [cell.value for cell in file_ws[1]]
        template_headers = [cell.value for cell in template_ws[1]]
        mismatch = [h for h in file_headers if h not in template_headers]
        
        if mismatch:
            header_mismatch[sheet] = mismatch
            
    return header_mismatch, sheet_mismatch

def quarantine_file(file_path, quarantined_path, file_error):
    file_name = os.path.basename(file_path)
    quarantine_file_path = os.path.join(quarantined_path, file_name)
    shutil.move(file_path, quarantine_file_path)
    
    with open(f"{quarantine_file_path}.txt", 'w') as error_log:
        error_log.write(file_error)

def create_log_entry(file_path, file_status, file_error):
    file_name = os.path.basename(file_path)
    created_time = datetime.fromtimestamp(os.path.getctime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
    scanned_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    return {
        'File Name': file_name,
        'Created Time': created_time,
        'Scanned Time': scanned_time,
        'Status': file_status,
        'Error': file_error
    }

def copy_successful_file(file_path, template_file, prod_path, archive_path):
    template_file_name = os.path.basename(template_file)
    prod_file_path = os.path.join(prod_path, template_file_name)
    archive_file_path = os.path.join(archive_path, f"{template_file_name}_{datetime.now().strftime('%Y%m%d_%H%M')}")
    
    shutil.copy2(file_path, prod_file_path)
    shutil.copy2(file_path, archive_file_path)

def generate_summary_report(summary_report, log_path):
    current_time = datetime.now().strftime('%Y%m%d_%H%M')
    txt_file = os.path.join(log_path, f"scan_{current_time}.txt")
    csv_file = os.path.join(log_path, f"scan_{current_time}.csv")
    
    with open(txt_file, 'w') as f:
        for entry in summary_report:
            f.write('\n'.join([f"{k}: {v}" for k, v in entry.items()]))
            f.write('\n\n')
            
    with open(csv_file, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=summary_report[0].keys())
        writer.writeheader()
        writer.writerows(summary_report)

if __name__ == '__main__':
    src_path = r'\\userserver\data'  # Set the source path here
    template_path = r'\\systemserver\data\templates'  # Set the template path here
    log_path = r'\\system'

if __name__ == '__main__':
    src_path = r'\\userserver\data'  # Set the source path here
    template_path = r'\\systemserver\data\templates'  # Set the template path here
    log_path = r'\\systemserver\logs'  # Set the log path here
    quarantined_path = r'\\systemserver\quarantined'  # Set the quarantined path here
    prod_path = r'\\systemserver\data\prod'  # Set the prod path here
    archive_path = r'\\systemserver\data\archive'  # Set the archive path here

    compare_files(src_path, template_path, log_path, quarantined_path, prod_path, archive_path)
